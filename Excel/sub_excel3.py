# author='lwz'
# coding:utf-8
# !/usr/bin/env python3
import openpyxl
from openpyxl import Workbook, load_workbook
import os
import time


Province_list1 = ['吉林', '辽宁', '北京', '天津', '上海',
                 '河北', '山西', '山东', '河南', '江苏', '安徽',
                 '浙江', '福建', '广东', '广西', '海南', '江西',
                 '湖北', '湖南', '云南', '贵州', '四川', '西藏',
                 '陕西', '重庆', '宁夏', '甘肃', '新疆', '青海',
                 ]
Province_list2 = ['黑龙江', '内蒙古']
Province_list = Province_list1 + Province_list2
City_list = ['北京', '天津', '上海', '重庆']
path = 'D:\\Cache\\Excel'


def read_excel():
    flag = 0
    start_time = time.time()
    workbook_list = list()
    workbook_name = list()
    sheet_list = list()
    #    打开文件 读取模式
    fn1 = r'C:\Users\lwzswufe\Documents\C 数据.xlsx'
    fn2 = r'C:\Users\lwzswufe\Documents\C.xlsx'
    workbook_now = load_workbook(fn1)
    #   打开指定sheet

    index = [0] * len(Province_list)

    for name in Province_list:
        fname = path + '\\' + name + '.xlsx'
        new_workbook = Workbook()  # 创建excel 得到workbook对象
        workbook_list.append(new_workbook)  # 记录workbook对象
        workbook_name.append(fname)
        # new_workbook.create_sheet("sheet")
        new_sheet = new_workbook.get_sheet_by_name("Sheet")
        sheet_list.append(new_sheet)
        print(fname)

    for sheet in workbook_now:
        if sheet.max_column == 6:
            idx = 4
        elif sheet.max_column == 5:
            idx = 3
        else:
            print(sheet.max_column)
            continue

        offset = 6 - sheet.max_column
        for i in range(1, sheet.max_row + 1):
            flag += 1
            province = sheet.cell(row=i, column=idx).value
            province_id = get_province_id(province)
            if province_id < 0:
                continue
            else:
                index[province_id] += 1
                row_flag = index[province_id]
                for j in range(1, sheet.max_column + 1):
                    sheet_list[province_id].cell(row=row_flag, column=offset + j).value \
                        = sheet.cell(row=i, column=j).value  # 写入每一列的数据

                if (flag + 1) % 100 == 0:
                    used_time = time.time() - start_time
                    print('{}rows used {:.4f}s'.format(flag+1, used_time))

    for i in range(len(workbook_list)):
        workbook_list[i].save(workbook_name[i])         # 保存workbook对象


def get_province_id(province_name):
    if len(province_name) <= 2:
        if province_name in City_list:
            return Province_list.index(province_name)
        else:
            print(province_name)
            return -1
    else:
        if province_name[:2] in Province_list:
            return Province_list.index(province_name[:2])
        elif province_name[:3] in Province_list2:
            return Province_list2.index(province_name[:3])
        else:
            print(province_name)
            return -1


if __name__ == '__main__':
    read_excel()
