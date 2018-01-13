# author='lwz'
# coding:utf-8
# !/usr/bin/env python3
import xlrd  # 打开模块
import xlwt  # 写入模块
import openpyxl
import os
import time


style = xlwt.XFStyle()
font = xlwt.Font()
font.name = 'SimSun'    # 指定“宋体”
font.height = 300  # 高度
font.bold = True  # 是否是粗体
style.font = font
Province_list1 = ['吉林', '辽宁', '北京', '天津', '上海',
                 '河北', '山西', '山东', '河南', '江苏', '安徽',
                 '浙江', '福建', '广东', '广西', '海南', '江西',
                 '湖北', '湖南', '云南', '贵州', '四川', '西藏',
                 '陕西', '重庆', '宁夏', '甘肃', '新疆', '青海',
                 ]
Province_list2 = ['黑龙江', '内蒙古']
Province_list = Province_list1 + Province_list2
City_list = ['北京', '天津', '上海', '重庆']
path = 'D:\\Cache\\Excel2'


def read_excel():
    flag = 0
    start_time = time.time()
    workbook_list = list()
    workbook_name = list()
    sheet_list = list()
    #    打开文件 读取模式
    workbook = xlrd.open_workbook(r'C:\Users\lwzswufe\Documents\C 数据.xlsx')
    #   打开指定sheet

    index = [0] * len(Province_list)

    for name in Province_list:
        fname = path + '\\' + name + '.xlsx'
        new_workbook = xlwt.Workbook(fname)  # 创建excel 得到workbook对象
        workbook_list.append(new_workbook)  # 记录workbook对象
        workbook_name.append(fname)
        new_workbook.add_sheet("sheet1")
        new_sheet = new_workbook.get_sheet(0)
        sheet_list.append(new_sheet)

    for sheet in workbook.sheets():
        if sheet.ncols == 6:
            idx = 3
        elif sheet.ncols == 5:
            idx = 2
        else:
            print(sheet.ncols)
            continue

        offset = 6 - sheet.ncols
        for i in range(1, min(sheet.nrows, 65500)):
            flag += 1
            province = sheet.cell(i, idx).value
            province_id = get_province_id(province)
            if province_id < 0:
                continue
            else:
                index[province_id] += 1
                row_flag = index[province_id]
                if row_flag >= 65500:
                    continue
                    
                for j in range(sheet.ncols):
                    sheet_list[province_id].write(row_flag, offset + j, sheet.cell(i, j).value)  # 写入每一列的数据

                if (flag + 1) % 100 == 0:
                    used_time = time.time() - start_time
                    print('{}rows used {:.4f}s'.format(flag, used_time))

    for i in range(len(workbook_list)):
        workbook_list[i].save(workbook_name[i])         # 保存workbook对象


def get_province_id(province_name):
    if len(province_name) <= 2:
        if province_name in City_list:
            return Province_list.index(province_name)
        else:
            print(province_name)
            return -1
    else:
        if province_name[:2] in Province_list:
            return Province_list.index(province_name[:2])
        elif province_name[:3] in Province_list2:
            return Province_list2.index(province_name[:3])
        else:
            print(province_name)
            return -1


if __name__ == '__main__':
    read_excel()
